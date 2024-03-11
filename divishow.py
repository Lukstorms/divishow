import openpyxl
from math import ceil
from openpyxl import Workbook
import os


def split_xlsx(filepath, number_of_files):
    # Load the workbook and active sheet
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active

    # Get the header from the first row
    headers = [cell.value for cell in sheet[1]]

    # Calculate the number of rows for each new file, excluding the header
    total_rows = sheet.max_row - 1
    rows_per_file = ceil(total_rows / number_of_files)

    # Initialize variables for splitting
    current_row = 2  # Start at the second row to skip the header
    for file_num in range(1, number_of_files + 1):
        # Create a new workbook and select the active sheet
        new_workbook = Workbook()
        new_sheet = new_workbook.active

        # Add the header to each new file
        new_sheet.append(headers)

        # Calculate the end row for the current file split
        end_row = min(current_row + rows_per_file, sheet.max_row + 1)

        # Copy rows to the new file
        for row_idx in range(current_row, end_row):
            row_data = [cell.value for cell in sheet[row_idx]]
            new_sheet.append(row_data)

        # Update the current row for the next file split
        current_row = end_row

        # Save the new file
        new_filename = f"{os.path.splitext(filepath)[0]}_part{file_num}.xlsx"
        new_workbook.save(new_filename)
        print(f"Created {new_filename}")


if __name__ == "__main__":
    filepath = input("Enter the filepath to the XLSX file: ").strip('"')
    number_of_files = int(input("Enter the number of files you want to create: "))
    split_xlsx(filepath, number_of_files)
