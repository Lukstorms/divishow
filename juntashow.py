import openpyxl
from openpyxl import Workbook
import os
import glob
import re


def merge_xlsx(filepaths, output_filepath):
    # Create a new workbook and select the active sheet
    merged_workbook = Workbook()
    merged_sheet = merged_workbook.active

    # Initialize variables for merging
    for i, filepath in enumerate(filepaths):
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active

        # Copy rows from each file to the merged sheet
        # Only copy the header from the first file
        for row in sheet.iter_rows(values_only=True):
            if i > 0 and row[0] == sheet['A1'].value:
                continue
            merged_sheet.append(row)

    # Save the merged file
    merged_workbook.save(output_filepath)
    print(f"Merged files saved to {output_filepath}")


def find_and_merge_files(base_name):
    # Use glob to find all files with the same base name
    filepaths = glob.glob(f"{base_name}_part*.xlsx")

    # Sort the files based on the part number
    filepaths.sort(key=lambda f: int(re.search(r'part(\d+)', f).group(1)))

    # Define the output file path in the current working directory instead of the root directory
    output_filepath = f"{base_name}_merged.xlsx"

    # Merge the files
    merge_xlsx(filepaths, output_filepath)


if __name__ == "__main__":
    base_name = input("Enter the base name of the XLSX files: ").strip('"')
    find_and_merge_files(base_name)
